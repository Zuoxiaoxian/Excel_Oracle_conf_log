# -*- coding: utf-8 -*-
# import sys
#
# reload(sys)
#
# sys.setdefaultencoding('utf8')


# 作者      ：xiaoxianzuo.zuo
# QQ        ：1980179070
# 文件名    ： excel_01.py
# 新建时间   ：2018/4/12/012 18:20



import os

import openpyxl

import re

cwd_path = os.getcwd()

parert_path = os.path.abspath(os.path.dirname(cwd_path) + os.path.sep + '.')

excel_path = os.path.join(parert_path, 'excel_workerbook')

# print(excel_path)
#example.xlsx需要位于当前工作目录中才能使用它,不是就要绝对路径！

# 解析指定文件夹下的Excel表！
def get_sheets():
    '''
    :return: [(),]
    '''
    _wb_sheetnames = []
    walks = os.walk(excel_path)
    try:
        for current_path, subfolders, filesnames in walks:
            for filename in filesnames:
                file_name = os.path.join(current_path, filename)
                # print file_name
                #F:\Python_Projects\py_excel\excel_workerbook\heart_rember.xlsx
                _wb = openpyxl.load_workbook(file_name)
                _sheetnames = _wb.sheetnames
                n = _wb, _sheetnames
                _wb_sheetnames.append(n)
                # for sheetname in sheetnames:
                #     sheet = wb[sheetname]
                #     print(sheet)
        print(_wb_sheetnames, '----------------------------------', type(_wb_sheetnames))
        # return _wb_sheetnames
    except Exception as e:
        print("错误！", e)
    return _wb_sheetnames

def re_search(strs):
    '''
    正则匹配
    :param strs: 要匹配的字符串
    :return: 元组， 行数， 列数
    '''
    pattern = re.compile(r'[A-Z]+')
    result = pattern.match(strs)
    result = result.group()
    result_index = strs.index(result)
    col_num = strs[result_index + 1:]
    row_num = result
    # print '正则匹配的result', strs, row_num, col_num
    return (row_num, col_num)


def row_or_col_nums(merge_split):
    '''
    合并的单元格，位置， 占的行数或列数
    :param merge_split: tupe合并的单元格
    :return: dict 占的行或列， 位置， 行数或列数
    '''
    row1, row2, col1, col2 = merge_split
    row_col_dict = dict()
    if row1 == row2:
        col_nums = int(col2) - int(col1)
        row_col = col_nums+1, row1+col1
        # 占的是列
        row_col_dict["row"] = row_col
    else:
        row_nums = ord(row2) - ord(row1)
        row_col = row1+col1, row_nums+1
        # 占的是行
        row_col_dict["col"] = row_col
        # print 'row_col_dict', row_col_dict
    return row_col_dict
merge_split_lists = []
rows_colums = []
# 默认行高、列宽
default_row_h = 20.1
default_col_w = 10
# def get_rowh_col_w(row_col, default_row_h, default_col_w):
#     if row_col.has_key("row"):
#         print default_row_h, default_col_w


def all_split_lists(max_row, max_column):
    '''
    :param max_row: 总行数
    :param max_column: 总列数
    :return: 如('J', 'L', '16', '16')
    '''
    max_row = int(max_row)
    max_column = int(max_column)
    rows = []
    colums = []
    for i in range(max_row):
        # print i + 1
        rows.append(str(i + 1))
    for j in range(max_column):
        # print chr(65 + j)
        colums.append(chr(65 + j))
    # print "行", rows
    # print "列", colums
    for i in rows:
        for j in colums:
            i_j = (j, str(int(i)))
            rows_colums.insert(int(i) - 1, i_j)
    rows_colums.sort(key=order_all_vars, reverse=False)
    # print rows_colums, '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
    return rows_colums


def unmergerd_cells(merge_split_lists, rows_colums):
    # print "合并的单元格： ", merge_split_lists
    # print "所有单元格： ", rows_colums
    for j in merge_split_lists:
        j_index = merge_split_lists.index(j)
        j1 = j[0]
        j2 = j[1]
        #'''移除合并行'''
        if j1 == j2:
            j3 = j[2]
            j4 = j[3]
            # print merge_split_lists[j_index], j3, j4
            for i in rows_colums:
                if i[0] == j1:
                    cha_zhi = int(j4) - int(j3)
                    if (i[0], j3) in rows_colums:
                        i_index = rows_colums.index((j1, j3))
                        # print rows_colums[i_index]
                        for k in range(cha_zhi + 1):
                            # print "移除合并行", rows_colums[i_index - k]
                            rows_colums.remove(rows_colums[i_index - k])
        # '''移除合并列'''
        else:
            j3 = j[2]
            cha_zhi = ord(j2) - ord(j1)
            # print (j1, j2, j3), cha_zhi, "||||||||||||||||||||||"
            for i in rows_colums:
                i_index2 = rows_colums.index(i)
                for l in range(cha_zhi + 1):
                    shan_chu = (chr(ord(j1) + l), j3)
                    if shan_chu in rows_colums:
                        # print "移除合并列：  ", shan_chu
                        rows_colums.remove(shan_chu)
    print( "---------------------------未合并的单元格------------------")
    # print "合并的单元格： ", merge_split_lists
    # print "没有合并的单元格 ： ", rows_colums
# def unmergerd_cells(merge_split_lists, rows_colums):
#     print "合并的单元格： ", merge_split_lists
#     print "所有单元格： ", rows_colums
#     for i in rows_colums:
#         i_index2 = rows_colums.index(i)
#         for j in merge_split_lists:
#             j_index = merge_split_lists.index(j)
#             j1 = j[0]
#             j2 = j[1]
#             #'''移除合并行'''
#             if j1 == j2:
#                 print "j1 == j2", j1, j2
#                 j3 = j[2]
#                 j4 = j[3]
#                 if i[0] == j1:
#                     # print "i[0], j1", i[0], j1, 'i', i
#                     cha_zhi = int(j4) - int(j3)
#                     if (i[0], j3) in rows_colums:
#                         i_index = rows_colums.index((i[0], j3))
#                     for k in range(cha_zhi + 1):
#                         print "移除合并行", rows_colums[i_index - k]
#                         rows_colums.remove(rows_colums[i_index - k])
#             #'''移除合并列'''
#             else:
#                 j3 = j[2]
#                 cha_zhi = ord(j2) - ord(j1)
#                 for l in range(cha_zhi):
#                     shan_chu = (chr(ord(rows_colums[i_index2][0]) + l + 1), j3)
#                     # print "移除合并列：  ", shan_chu
#                     if shan_chu in rows_colums:
#                         rows_colums.remove(shan_chu)
#     print "---------------------------未合并的单元格------------------"
#     print "合并的单元格： ", merge_split_lists
#     print "没有合并的单元格 ： ", rows_colums


def style_list(row_col):
    # print row_col
    if "row" in row_col:
        inner = row_col["row"]
        if isinstance(inner[0], int):
            inner = inner[1]
        else:
            inner = inner[0]
    else:
        inner = row_col["col"]
        if isinstance(inner[0], int):
            inner = inner[1]
        else:
            inner = inner[0]
    return inner


def thin_none(*args):
    thin_list = []
    for thin_none in args:
        if thin_none == None:
            thin_none = 0
        elif thin_none == "thin":
            thin_none = 1
        thin_list.append(thin_none)
    return thin_list



def order_hang(key):
    '''---------行排序--------'''
    # print key[2], '+++++++++++++'
    return int(key[2])


def order_hang2(key):
    '''---------行排序--------'''
    # print key[2], '+++++++++++++'
    return ord(key[0][0])


def order_lie(key):
    '''---------行排序--------'''
    # print key[1], '+++++++++++++'
    return key[1]


def order_all_vars(key):
    '''对全部表单的列表排序，按照列'''
    lie = ord(key[0])
    return key[0]


def list_order(merge_split_lists):
    '''
    首先按照 行进行排序，然后是列！
    :param strs:
    :return:
    '''
    merge_split_lists.sort(key=order_hang)
    merge_split_lists.sort(key=order_lie)
    # print len(merge_split_lists), "长度"
    # print merge_split_lists, "IIIIIIIIIIIIIIII"
    # return merge_split_lists, "IIIIIIIIIIIIIIII"


def get_inner():
    '''
    合并的单元格，
    top_right_bottom_left：获取表格，边框（上，右，左，中 顺时针 -- 0 没有边框）
    row_h_and_col_w： 行高，列宽， # None说明是默认值
    col： 位置， 占几列
    row： 位置， 占几行
    inner： 内容
    :return: [{}*****]
    '''
    _wb_sheetnames = get_sheets()
    # wb, sheetnames = get_sheets()
    wb, sheetnames = _wb_sheetnames[0]
    sheet = wb[sheetnames[0]]
    # print "&&&&&&&&&&&&&&&&&&&&&&&&&&", dir(sheet)
    style_result_list = []
    ''' 得到默认行高， 和列宽  '''
    # max_row = int(sheet.max_row) + 1
    # max_column = int(sheet.max_column) + 1
    # default_row_h = sheet.row_dimensions[max_row].height
    # default_col_w = sheet.column_dimensions[chr(65 + max_column)].width

    # 行列
    # print sheet.max_row, sheet.max_column
    max_row, max_column = sheet.max_row, sheet.max_column
    '''--------------------得到表单的行数、列数进行处理获取内容位置类似('J', 'L', '16', '16')-------------------------'''
    all_split_lists(max_row, max_column)
    print( "sheet.merged_cells:   ", str(sheet.merged_cells))

    for i in sheet.merged_cells:
        # print i, type(i)
        merge = str(i)
        # print(merge, '--------------------------')
        merge = merge.split(":")
        # print(merge, '--------------------------')
        start, end = merge[0], merge[1]
        row_num, col_num = re_search(start)
        row_num2, col_num2 = re_search(end)
        merge_split = (row_num, row_num2, col_num, col_num2)
        # print merge_split, "PPPPPPPPPPPPPPPPPPP"
        merge_split_lists.append(merge_split)
        # print merge_split, "PPPPPPPPPPPP"
        # row_col = row_or_col_nums(merge_split)
        # print row_col, 'OOOOOOOOOOOOOOOO'
        # '''-----------------行高， 列宽取值--------------------'''
        # # sheet.row_dimensions[6].height, sheet.column_dimensions['A'].width
        # # row_col, default_row_h, default_col_w
        # row_h_list = col_w_list = []
        # if row_col.has_key("row"):
        #     _row_num = row_col.get("row")[0]
        #     _row_position = row_col.get("row")[1]
        #     row_h = sheet.row_dimensions[int(re_search(_row_position)[1])].height
        #     col_w = sheet.column_dimensions[re_search(_row_position)[0]].width
        #     for i in range(int(re_search(_row_position)[1]), _row_num + int(re_search(_row_position)[1])):
        #         row_h_list.append(sheet.row_dimensions[i].height)
        #     row_h_finaly = sum(row_h_list)
        #     col_w_finaly = default_col_w
        #     # print row_h_list, row_h_finaly, 'ppppp'
        # else:
        #     _col_num = row_col.get("col")[1]
        #     _col_position = row_col.get("col")[0]
        #     # print _col_num, _col_position
        #     # print re_search(_col_position)[0], int(re_search(_col_position)[1])
        #     ii = re_search(_col_position)[0]
        #     for i in range(int(re_search(_col_position)[1]), int(re_search(_col_position)[1]) + _col_num):
        #         # print ii
        #         if sheet.column_dimensions[ii].width == None:
        #             sheet.column_dimensions[ii].width = default_col_w
        #             col_w_list.append(sheet.column_dimensions[ii].width)
        #         else:
        #             col_w_list.append(sheet.column_dimensions[ii].width)
        #         ii = chr(ord(re_search(_col_position)[0]) + i)
        #     col_w_finaly = sum(col_w_list)
        #     row_h_finaly = default_row_h
        # inner = style_list(row_col)
        # inner_inner = sheet[inner].value
        # row_col["inner"] = inner_inner
        # # style_result_list.append(row_col)
        # '''--------------表格边距------------------'''
        # #'top', 'right', 'bottom', 'left'
        # border = sheet[inner].border
        # inner_top, inner_right, inner_bottom, inner_left = border.top.style, border.right.style, border.bottom.style, border.left.style
        # thin_list = thin_none(inner_top, inner_right, inner_bottom, inner_left)
        #
        # # row_col["top_right_bottom_left"] = inner_top, inner_right, inner_bottom, inner_left
        # row_col["top_right_bottom_left"] = thin_list
        # # print "inner_top : %s, inner_right : %s, inner_bottom : %s, inner_left : %s, inner : %s " % (inner_top, inner_right, inner_bottom, inner_left, inner)
        # '''
        # 如果， border1.left.style 是None ---> 说明单元格左边框为： 没有
        # 如果， border1.left.style 是thin ---> 说明单元格左边框为： 一般设置
        # '''
        # style_result_list.append(row_col)
        # '''--------------表格边距------------------'''
        # # row_dimensions, column_dimensions ； 一行的高度， 一列的宽度
        # #width
        # # print sheet.column_dimensions['A'], '<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>'
        # # print sheet.row_dimensions[3], '<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>', dir(sheet.row_dimensions[3])
        # #10.625 20.1
        # # print sheet.row_dimensions[6].height, sheet.column_dimensions['A'].width
        # '''----------------表格宽高-----------------'''
        # row_col["row_h_and_col_w"] = row_h_finaly, col_w_finaly
    '''------------------------得到没有合并过的单元格--------------------------------'''
    unmergerd_cells(merge_split_lists, rows_colums)
    '''---------------------- 排序！ -----------------------'''
    # print "合并的单元格： ", merge_split_lists
    list_order(merge_split_lists)
    ########################未合并的单元格##################################
    for merge_split in merge_split_lists:
        # print merge_split, "__________________"
        row_col = row_or_col_nums(merge_split)
        # print row_col, 'OOOOOOOOOOOOOOOO'
        '''-----------------行高， 列宽取值--------------------'''
        # sheet.row_dimensions[6].height, sheet.column_dimensions['A'].width
        # row_col, default_row_h, default_col_w
        row_h_list = []
        col_w_list = []
        if "row" in  row_col:
            _row_num = row_col.get("row")[0]
            _row_position = row_col.get("row")[1]
            row_h = sheet.row_dimensions[int(re_search(_row_position)[1])].height
            col_w = sheet.column_dimensions[re_search(_row_position)[0]].width
            for i in range(int(re_search(_row_position)[1]), _row_num + int(re_search(_row_position)[1])):
                # print "i", i
                row_h_list.append(sheet.row_dimensions[i].height)
            row_h_finaly = sum(row_h_list)
            col_w_finaly = default_col_w
            # print row_h_list, row_h_finaly, 'ppppp'
        else:
            _col_num = row_col.get("col")[1]
            _col_position = row_col.get("col")[0]
            # print _col_num, _col_position
            # print re_search(_col_position)[0], int(re_search(_col_position)[1])
            ii = re_search(_col_position)[0]
            for i in range(int(re_search(_col_position)[1]), int(re_search(_col_position)[1]) + _col_num):
                # print ii, "ii"
                if sheet.column_dimensions[ii].width == None:
                    sheet.column_dimensions[ii].width = default_col_w
                    col_w_list.append(sheet.column_dimensions[ii].width)
                else:
                    col_w_list.append(sheet.column_dimensions[ii].width)
                ii = chr(ord(re_search(_col_position)[0]) + i)
            col_w_finaly = sum(col_w_list)
            row_h_finaly = default_row_h
        inner = style_list(row_col)
        print("inner", inner)
        inner_inner = sheet[inner].value
        # print "E4:::::", sheet["E4"].border
        # print "E6:::::", sheet["E6"].border
        # print '-----'
        # print "F3:::::", sheet["F3"].border
        # print "E3:::::", sheet["E3"].border

        row_col["inner"] = inner_inner
        # style_result_list.append(row_col)
        '''--------------表格边距------------------'''
        # 'top', 'right', 'bottom', 'left'
        border = sheet[inner].border
        inner_top, inner_right, inner_bottom, inner_left = border.top.style, border.right.style, border.bottom.style, border.left.style
        thin_list = thin_none(inner_top, inner_right, inner_bottom, inner_left)
        # row_col["top_right_bottom_left"] = inner_top, inner_right, inner_bottom, inner_left
        row_col["top_right_bottom_left"] = thin_list
        # print "inner_top : %s, inner_right : %s, inner_bottom : %s, inner_left : %s, inner : %s " % (inner_top, inner_right, inner_bottom, inner_left, inner)
        '''
        如果， border1.left.style 是None ---> 说明单元格左边框为： 没有
        如果， border1.left.style 是thin ---> 说明单元格左边框为： 一般设置
        '''
        style_result_list.append(row_col)
        '''--------------表格边距------------------'''
        # row_dimensions, column_dimensions ； 一行的高度， 一列的宽度
        # width
        # print sheet.column_dimensions['A'], '<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>'
        # print sheet.row_dimensions[3], '<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>', dir(sheet.row_dimensions[3])
        # 10.625 20.1
        # print sheet.row_dimensions[6].height, sheet.column_dimensions['A'].width
        '''----------------表格宽高-----------------'''
        row_col["row_h_and_col_w"] = row_h_finaly, col_w_finaly
    return style_result_list
    # "最大行数"
    #max_row = sheet.max_row
    # "最大列数"
    #max_column = sheet.max_column
    # print sheet["A18"].value
    # print chr(65 + int(max_column) -1)
    # print sheet.column_dimensions, '////////////////////'
    # print dir(sheet)
    # for i in sheet.columns:
    #     print i, '---------'
    # for row in range(int(max_row)):
    #     hang_inner = dict()
    #     hang_inner_list = []
    #     for col in range(int(max_column)):
    #         cell = sheet["%s%d" % (chr(65 + col), row + 1)].value
    #         # print 'cell', cell, "%s%d" % (chr(65 + col), row + 1)
    #         lie_char = chr((65 + col))
    #         hang_int = row + 1
    #         present = "%s%d" % (lie_char, hang_int)
    #         hang_inner_list.append((cell, present))
    #     hang_inner[str(row + 1) + 'row'] = hang_inner_list
    #     # print hang_inner_list
    #     # for inner in range(len(hang_inner_list)):
    #     #     print hang_inner_list[inner]
    #     #     start_p = end_p = inner
    #     #     if hang_inner_list[inner] != None:
    #     #         while inner < len(hang_inner_list)-1:
    #     #             inner += 1
    #     #             print inner, hang_inner_list[inner]
    #     #             if hang_inner_list[inner] != None:
    #     #                 inner_next = inner
    #     #                 end_p = inner_next
    #     #                 print hang_inner_list[inner_next], inner_next
    #     #
    #     #                 # input("----------")
    #     #             else:
    #     #                 continue
    #
    #
    #
    #     # print '>>>>>>>>>>><<<<<<<<<<'
    #     # print hang_inner
    #     cell_style = hang_inner_list[0]
    #     print hang_inner
    # # return hang_inner
    # return style_result_list

    '''
    for sheetname in sheetnames:
        sheet = wb[sheetname]
        print '当前的sheet页名： ', sheet
        cell = sheet['A1']
        
        print "cell.value", cell.value
        style_names = wb.style_names
        print 'style_names', style_names[0]

        print 'inner style:', cell.style
#  cell.font, cell.fill
        print '##############', cell.has_style
        max_row = sheet.max_row
        max_column = sheet.max_column
        # "最大行数"
        print "最大行数",max_row
        # "最大列数"
        print "最大列数", max_column
        print sheet.row_dimensions[1].height, '行高'
        print sheet.column_dimensions['B'].width, '列宽'
        return sheet
    '''


def get_inner_unmerger():
    _wb_sheetnames = get_sheets()
    rows_colums_style = []
    # wb, sheetnames = get_sheets()[0]
    wb, sheetnames = _wb_sheetnames[0]

    sheet = wb[sheetnames[0]]
    for i, row_colum in enumerate(rows_colums):
        # 单个表格值
        inner = row_colum[0] + row_colum[1]
        value = sheet[inner].value
        zuo_baio = (row_colum[0], row_colum[1])
        '''单个表格： 开始行， 结束行，占几行，占几列'''
        rstart = rend = int(row_colum[1])
        erow = ecol = 1

        if value == None:
            value = 'null'
            isedit = 1
        else:
            isedit = 0
        #单个表格边距,有无变框
        border = sheet[inner].border
        inner_top, inner_right, inner_bottom, inner_left = border.top.style, border.right.style, border.bottom.style, border.left.style
        thin_list = thin_none(inner_top, inner_right, inner_bottom, inner_left)
        bold_value = str(thin_list[0]) + str(thin_list[1]) + str(thin_list[2]) + str(thin_list[3])
        row_h = sheet.row_dimensions[row_colum[1]].height
        colum_w = sheet.column_dimensions[row_colum[0]].width
        if row_h == None:
            row_h = default_row_h
        if colum_w == None:
            colum_w = default_col_w
        rows_colums_style.insert(i, (zuo_baio, value, rstart, rend, erow, ecol, colum_w, row_h, isedit, bold_value))
    # print "合并的单元格： ", merge_split_lists
    # print "没有合并的单元格 ： ", rows_colums
    return rows_colums_style
    '''----------------没有合并的单元格解析------------------------'''


def parsre_list_unmerger():
    '''为合并的项'''
    rows_colums_style = get_inner_unmerger()
    rows_colums_style.sort(key=order_hang)
    rows_colums_style.sort(key=order_hang2)
    # for i in rows_colums_style:
    #     print i
    return rows_colums_style


def parsre_list():
    style_result_list = get_inner()
    # print "style_result_list", style_result_list
    result_list = []
    for o, style_result in enumerate(style_result_list):
        # result = []
        # VALUE
        value = style_result["inner"]
        if value:
            isedit = 0
        else:
            isedit = 1
            value = 'null'
        if "col" in style_result:
            print( "col", style_result["col"])
            rstart = style_result["col"][0]
            # 正则，
            pattern = re.compile(r'[A-Z]+')
            result = pattern.match(rstart[0])
            result = result.group()
            result_index = rstart[0].index(result)
            zuo_biao = (result, rstart[result_index + 1: ])
            # print "zuo_baio", zuo_baio
            ecol = style_result["col"][1]
            rstart = rend = int(rstart[result_index + 1: ])
            erow = 0
            # print("开始行: ", rstart, "占几列: ", ecol, "结束行: ", rend)
            # print "开始行: ", rstart, "占几行: ", erow, "结束列: ", rend
            # print (style_result["col"], "列")
        else:
            print( "row", style_result["row"])
            lstart = style_result["row"][1]
            erow = style_result["row"][0]
            # 正则
            pattern = re.compile(r'[A-Z]+')
            result = pattern.match(lstart[0])
            result = result.group()
            result_index = lstart.index(result)
            zuo_biao = (result, lstart[result_index + 1: ])
            # print lstart[result_index + 1: ], "lstart[-1:result_index]", type(result_index), result_index
            rstart = int(lstart[result_index + 1: ])
            rend = rstart + erow - 1
            ecol = 0
            # print "开始行: ", rstart, "占几行: ", erow, "结束列: ", rend
            # print style_result["row"], "行"

        height = int(style_result['row_h_and_col_w'][0])
        width = int(style_result['row_h_and_col_w'][1])
        # isedit = None
        bold = [str(i) for i in style_result['top_right_bottom_left']]
        bold = bold[0] + bold[1] + bold[2] + bold[3]
        results = (zuo_biao, value, rstart, rend, erow, ecol, width, height, isedit, bold)
        # result.extend(results)
        result_list.insert(o, results)
    print( "合并的项，", result_list)
    return result_list


def start():
    # 合并的
    result_list = parsre_list()
    print('合并的:', result_list)
    # 未合并的
    rows_colums_style = parsre_list_unmerger()
    print("rows_colums_style：", rows_colums_style)
    zong_list = result_list + rows_colums_style
    zong_list.sort(key=order_hang)
    zong_list.sort(key=order_hang2)
    for zong in zong_list:
        print( zong)
    return zong_list

if __name__ == '__main__':
    start()

    # style_result_list = get_inner()
    # print style_result_list

    # get_sheets()
    # print sheet.row_dimensions[1].height, '行高'
