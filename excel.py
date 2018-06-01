# -*- coding: utf-8 -*-
# 作者      ：xiaoxianzuo.zuo
# QQ        ：1980179070
# 文件名    ： excel_01.py
# 新建时间   ：2018/4/12/012 18:20



import os

import openpyxl

import re


#example.xlsx需要位于当前工作目录中才能使用它,不是就要绝对路径！

# 默认行高、列宽
# default_row_h = 20
# default_col_w = 10


class ParseSheetZxx(object):
    def __init__(self, excel_path, default_row_h=20, default_col_w=10):
        self.excel_path = excel_path
        self.default_row_h = default_row_h
        self.default_col_w = default_col_w
    '''# 解析指定文件夹下的Excel表！'''
    def get_sheets(self):
        '''
        :return: [(),]
        '''
        self._wb_sheetnames = []
        walks = os.walk(self.excel_path)
        #<generator object walk at 0x000002458B6198E0>
        try:
            for current_path, subfolders, filesnames in walks:
                for filename in filesnames:
                    file_name = os.path.join(current_path, filename)
                    # file_name = F:\Python_Projects\py_excel\excel_workerbook\heart_rember.xlsx
                    _wb = openpyxl.load_workbook(file_name)
                    _sheetnames = _wb.sheetnames
                    n = _wb, _sheetnames
                    self._wb_sheetnames.append(n)
        except Exception as e:
            print("错误！", e)
        wb_sheetnames = self._wb_sheetnames
        return wb_sheetnames

    '''正则解析， 将数字与字母分开'''
    def re_parse(self, start, end=None):
        pattern = re.compile(r'[A-Z]+')
        if end == None:
            result = pattern.match(start)
            result = result.group()
            result_index = start.index(result)
            start_str, start_num = result, start[result_index + 1:]
            start = start_str, start_num
            return start
        else:
            result = pattern.match(start)
            result = result.group()
            result_index = start.index(result)
            start_str, start_num = result, start[result_index + 1:]
            start = (start_str, start_num)

            result_end = pattern.match(end)
            result_end = result_end.group()
            result_index = end.index(result_end)
            end_str, end_num = result_end, end[result_index + 1:]
            end = (end_str, end_num)
            return start, end

    '''解析合并。。。得到：[[('J', 'L', '1', '1'),('A', 'A', '2', '2'), ], [('J', 'L', '1', '1'), ('A', 'A', '2', '2'), ]]'''
    def parse_merger(self, mergers_lists):
        '''
        :param mergers_lists:合并项，[<MultiCellRange [J16:L16 A20:A24],<MultiCellRange [J16:L16 A20:A24]>]
        :return:[[('J', '1', 'L', '1'),('A', '2', 'A', '2'), ], [('J', '1', 'L', '1'), ('A', '2', 'A', '2'), ]]
        '''
        # print(mergers_lists)
        start_end_list_all = []
        for merger_list in mergers_lists:
            start_end_list = []
            for merger in merger_list:
                merger = str(merger).split(":")
                start, end = merger[0], merger[1]
                # 正则解析，将字母与数字分开！
                start, end = self.re_parse(start, end)
                start_end = (start[0], end[0], start[1], end[1])
                start_end_list.append(start_end)
            start_end_list = list(set(start_end_list))
            start_end_list_all.append(start_end_list)
        # print("start_end_list_all", len(start_end_list_all), start_end_list_all)
        return start_end_list_all

    '''解析全部。。。得到：[[('A', '1'), ('A', '2')], [('A', '1'), ('A', '2')]]'''
    def parse_all(self, max_row_max_column):
        start_end_all_lists = []
        for row_column in max_row_max_column:
            start_end_list = []
            row, column = row_column[0], row_column[1]
            for c in range(column):
                column_chr = chr(c + 65)
                for r in range(row):
                    start_end = (column_chr, str(r + 1))
                    start_end_list.append(start_end)
            start_end_all_lists.append(start_end_list)
        return start_end_all_lists

    '''根据全部与和合并项，得到没有合并的项！'''
    def unmerger(self, start_end_list_all, start_end_all_lists):
        # print("解析合并-位置", start_end_list_all)
        # print("解析全部-位置", start_end_all_lists)
        for i in range(len(start_end_all_lists)):
            for start_end_merger in start_end_list_all[i]:
                # print("合并-位置: ", start_end_merger)
                # print("全部-位置: ", start_end_all)
                merger_weizhi0 = start_end_merger[0]
                merger_weizhi1 = start_end_merger[1]
                # 移除合并行
                if merger_weizhi0 == merger_weizhi1:
                    merger_weizhi2 = start_end_merger[2]
                    merger_weizhi3 = start_end_merger[3]
                    for start_end_all in start_end_all_lists[i]:
                        if start_end_all[0] == merger_weizhi0:
                            cha_zhi = int(merger_weizhi3) - int(merger_weizhi2)
                            sha_chu = (start_end_all[0], merger_weizhi2)
                            if sha_chu in start_end_all_lists[i]:
                                mergen_index = start_end_all_lists[i].index(sha_chu)
                                # print("mergen_index", mergen_index)
                                for k in range(cha_zhi + 1):
                                    start_end_all_lists[i].remove(start_end_all_lists[i][mergen_index])
                # 移除合并列
                else:
                    merger_weizhi2 = start_end_merger[2]
                    cha_zhi = ord(start_end_merger[1]) - ord(start_end_merger[0])
                    for l in range(cha_zhi + 1):
                        shan_chu = (chr(ord(merger_weizhi0) + l), merger_weizhi2)
                        # print("删除的合并列： ", shan_chu)
                        if shan_chu in start_end_all_lists[i]:
                            # print("删除！")
                            start_end_all_lists[i].remove(shan_chu)
                    # print(" 没有合并的列：", start_end_all_lists[i])
                # print("最终：  ", start_end_all_lists[i])
        # print("移除合并行-移除合并列", start_end_all_lists)
        return start_end_all_lists

    '''解析表格的边框样式， 主要是是否有边框！'''
    def is_or_thin(self, *args):
        is_or_thin = []
        bold_value = ''
        for thin_none in args:
            if thin_none == None:
                thin_none = '0'
            elif thin_none == "thin":
                thin_none = '1'
            is_or_thin.append(thin_none)
        for i in is_or_thin:
            bold_value += i
        return bold_value

    '''最终解析没有合并的，得到指定格式！
       得到： 如，[(('A', '24'), '均布8点 \n跳动值', 6, 15, 10, 0, 10, 200, 0, '1101'),]
       wei_zhi, value, rstart, rend, erow, ecol, width, height, isedit, bold_value
       [[((wei_zhi), alue, rstart, rend, erow, ecol, width, height, isedit, bold_value), (),()]]
    '''
    def end_parse_unmerger_lists(self, unmerger_lists):
        end_parse_unmerger_lists = []
        # print("len(unmerger_lists)", len(unmerger_lists), unmerger_lists)
        for i in range(len(unmerger_lists)):
            end_parse_unmerger_list = []
            unmerger_list = unmerger_lists[i]
            sheet = self.sheet_list[i]
            for end_parse in unmerger_list:
                wei_zhi = end_parse
                inner = wei_zhi[0] + wei_zhi[1]
                value = sheet[inner].value
                if value == None:
                    value = 'null'
                    isedit = 1
                else:
                    isedit = 0
                rend = rstart = int(wei_zhi[1])
                erow = ecol = 1
                width, height = self.default_col_w, self.default_row_h
                border = sheet[inner].border
                inner_top, inner_right, inner_bottom, inner_left = border.top.style, border.right.style, border.bottom.style, border.left.style
                bold_value = self.is_or_thin(inner_top, inner_right, inner_bottom, inner_left)
                # print("!!!!!!!!!!!!!!!:", wei_zhi, value, rstart, rend, erow, ecol, width, height,  isedit, bold_value)
                end_parse_unmerger_list.append([wei_zhi, value, rstart, rend, erow, ecol, width, height,  isedit, bold_value])
            end_parse_unmerger_lists.append(end_parse_unmerger_list)
        return end_parse_unmerger_lists

    '''最终解析 合并了的项！wei_zhi, value, rstart, rend, erow, ecol, width, height, isedit, bold_value'''
    def end_parse_merger_lists(self, start_end_list_all):
        end_parse_merger_lists = []
        for i in range(len(start_end_list_all)):
            end_parse_merger_list = []
            merger_list = start_end_list_all[i]
            sheet = self.sheet_list[i]
            for merger in merger_list:
                wei_zhi = (merger[0], merger[2])
                inner = merger[0] + merger[2]
                value = sheet[inner].value
                if value == None:
                    value = 'null'
                    isedit = 1
                else:
                    isedit = 0
                rstart = int(merger[2])
                rend = int(merger[3])
                # 解析合并项的 宽高！行高：sheet.row_dimensions[6].height,列宽：sheet.column_dimensions['A'].width
                if merger[0] == merger[1]:
                    erow = rend - rstart
                    ecol = 1
                    width = int(sheet.column_dimensions[merger[0]].width)
                    height = 0
                    for r in range(rstart, rend + 1):
                        height += sheet.row_dimensions[r].height
                        self.height = int(height)
                        inner = merger[0] + str(r)
                        # print("=====", inner) #A7
                        # 表格边框！
                        border = sheet[inner].border
                        inner_top, inner_right, inner_bottom, inner_left = border.top.style, border.right.style, border.bottom.style, border.left.style
                        bold_value = self.is_or_thin(inner_top, inner_right, inner_bottom, inner_left)
                    end_parse_merger_list.append([wei_zhi, value, rstart, rend, erow, ecol, width, 'height', isedit, bold_value])
                    # print("height", height)
                    # print("end_parse_merger_list", end_parse_merger_list)
                else:
                    erow = 1
                    ecol = ord(merger[1]) - ord(merger[0]) + 1
                    height = int(sheet.row_dimensions[rstart].height)
                    width_ = 0
                    for e in range(ecol):
                        e = chr(ord(merger[0]) + e)
                        width = sheet.column_dimensions[e].width
                        if width == None:
                            width = self.default_col_w
                        width_ += width
                        inner = e + merger[2]
                        # print("=====", inner)
                        # 表格边框！
                        border = sheet[inner].border
                        inner_top, inner_right, inner_bottom, inner_left = border.top.style, border.right.style, border.bottom.style, border.left.style
                        bold_value = self.is_or_thin(inner_top, inner_right, inner_bottom, inner_left)
                    end_parse_merger_list.append([wei_zhi, value, rstart, rend, erow, ecol, 'width', height, isedit,bold_value])
                    width = int(width_)
                    self.width = width
                    # print("end_parse_merger_list", len(end_parse_merger_list))
                    # end_parse_merger_list[6] = width
            for i in end_parse_merger_list:
                if i[6] == 'width':
                    i[6] = self.width
                if i[7] == 'height':
                    i[7] = self.height
            end_parse_merger_lists.append(end_parse_merger_list)
                # print("合并：wei_zhi, value, rstart, rend, erow, ecol, width, height, isedit", wei_zhi, value, rstart, rend, erow, ecol, width, height, isedit)
        # print("最终解析的，合并： ", end_parse_merger_lists)
        return end_parse_merger_lists

    '''定义排序规则！'''
    def order_list_lie(self, key):
        # print(key[0], '>>>>>', key[1])
        return key[0][0]

    def order_list_row(self, key):
        return int(key[0][1])

    ''' 排序：将数据进一步排序！ '''
    def dict_orser(self, zong_lists):
        # 排序：将数据进一步排序！
        for zong_list in zong_lists:
            zong_list.sort(
                key=lambda x: (int(x[0][1]), x[0][0])
            )
        # print("zong_lists", zong_lists)
        return zong_lists
        #     zong_list2 = []
        #     for zong in zong_list:
        #         zong_list2.append([zong[0][0], zong[0][1], zong[1: ]])
        #     print("zong_list2", zong_list2)
        #     zong_list2.sort(
        #         key=lambda x: (int(x[1]), x[0])
        #     )
        #     zong_lists2.append(zong_list2)
        #
        #     # zong_lists2.append(zong_list2)
        # print("zong_lists2", zong_lists2)
        # zong_lists = zong_lists2
        # return zong_lists2



    ''' 排序：将合并的与没有合并的进行，排序  '''
    def merger_unmerger_order(self, end_parse_unmerger_lists, end_parse_merger_lists):
        zong_lists = []
        for i in range(len(end_parse_unmerger_lists)):
            end_parse_unmerger_list = end_parse_unmerger_lists[i]
            end_parse_merger_list = end_parse_merger_lists[i]
            merger_unmerger_list = end_parse_unmerger_list + end_parse_merger_list
            # 排序。。。。。。
            merger_unmerger_list.sort(key=self.order_list_row)
            # merger_unmerger_list.sort(key=self.order_list_lie)
            # for i in merger_unmerger_list:
            #     print(i)
            zong_lists.append(merger_unmerger_list)
        return zong_lists

    '''得到sheet表单[(总的行数、列数)]，[合并项]'''
    def get_merger_and_all(self, wb_sheetnames):
        # (zuo_biao, value, rstart, rend, erow, ecol, width, height, isedit, bold)
        mergers_lists = []
        max_row_max_column = []
        wb_sheetnames = wb_sheetnames
        self.sheet_list = []
        for wb_sheetname in wb_sheetnames:
            wb, sheetnames = wb_sheetname
            # 第一个sheet
            sheet = wb[sheetnames[0]]
            self.sheet_list.append(sheet)
            # 和并项的坐标！
            mergers = sheet.merged_cells
            mergers_lists.append(mergers)
            # 表单的总行数、列数！
            max_row, max_column = sheet.max_row, sheet.max_column
            max_row_max_column.append((max_row, max_column))
        # return max_row_max_column, mergers_lists

        # 解析合并-位置！
        start_end_list_all = self.parse_merger(mergers_lists)

        # 解析全部-位置！
        start_end_all_lists = self.parse_all(max_row_max_column)

        # 根据合并的与全部的，得到没有合并的！
        unmerger_lists = self.unmerger(start_end_list_all, start_end_all_lists)

        # 最终解析没有合并的项！
        end_parse_unmerger_lists = self.end_parse_unmerger_lists(unmerger_lists)
        # for i in end_parse_unmerger_lists:
        #     print("end_parse_unmerger_lists", end_parse_unmerger_lists)
        #     input("******************************")
        # 最终解析合并的项！
        end_parse_merger_lists = self.end_parse_merger_lists(start_end_list_all)

        # list 排序：将合并的与没有合并的进行，排序
        zong_lists = self.merger_unmerger_order(end_parse_unmerger_lists, end_parse_merger_lists)
        # list 生成 dict 排序：将数据进一步排序！
        zong_lists = self.dict_orser(zong_lists)
        # print("一步排序  后：", zong_lists)
        return zong_lists

'''主函数'''
def excel_01(excel_path):
    parse_sheet = ParseSheetZxx(excel_path)
    wb_sheetnames = parse_sheet.get_sheets()
    zong_lists = parse_sheet.get_merger_and_all(wb_sheetnames)
    print("一步排序  后：", zong_lists)
    return zong_lists

if __name__ == '__main__':
    pass
    # parse_sheet = ParseSheet()
    # wb_sheetnames = parse_sheet.get_sheets()
    # max_row_max_column, mergers_lists = parse_sheet.get_merger_and_all(wb_sheetnames)
    # print(max_row_max_column, mergers_lists)


    # cwd_path = os.getcwd()
    # parert_path = os.path.abspath(os.path.dirname(cwd_path) + os.path.sep + '.')
    # excel_path = os.path.join(parert_path, 'excel_workerbook')
    # print("excel_path", excel_path)
    # excel_01(excel_path)
